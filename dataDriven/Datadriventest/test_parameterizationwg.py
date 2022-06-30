import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from dataDriven.XLUtils import XLutils
import time

s = Service(executable_path="/Users/venkateshr/PycharmProjects/WeightGurus/drivers/chromedriver")

class TestWeightgurus():
    @pytest.fixture()
    def setup(self):
        self.driver = webdriver.Chrome(service=s)
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        yield
        self.driver.close()
        self.driver.quit()

    def test_readData(setup):
        list=[]
        path = "/Users/venkateshr/Desktop/datadrivenwglogin.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name('Sheet1')

        rows = sheet.max_row
        cols = sheet.max_column

        print(rows)
        print(cols)

        for r in range(2, rows + 1):
            #for c in range(1, cols + 1):
             #   print(sheet.cell(r, c).value, end='  ')
            #print('\n')
            username = sheet.cell(r, 1).value
            password = sheet.cell(r, 2).value
            tuple=(username,password)
            list.append(tuple)
            print(list)
            return list

    @pytest.mark.parametrize("username, password",test_readData())
    def test_login(self,setup,username,password):
        self.driver.get("http://192.168.0.118:8100/landing")
        time.sleep(3)
        self.driver.find_element(By.XPATH, "(//ion-button[@shape='round'])[1]").click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//ion-input[@type='email']/input").send_keys(username)
        self.driver.find_element(By.XPATH, "//input[@name='ion-input-1']").send_keys(password)
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//ion-button[@size='large']").click()
        time.sleep(4)





