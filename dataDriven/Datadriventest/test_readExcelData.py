import openpyxl

def test_readData():
    path = "/Users/venkateshr/Desktop/datadrivenwglogin.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name('Sheet1')

    rows = sheet.max_row
    cols = sheet.max_column

    print(rows)
    print(cols)

    for r in range(2,rows+1):
        for c in range(1,cols+1):
            print(sheet.cell(r, c).value,end='  ')
        print('\n')




